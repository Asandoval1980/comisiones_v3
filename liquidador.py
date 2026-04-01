from __future__ import annotations

import copy
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

TRM_PESOS = 3348
TC_PAYMENT_METHODS = {"TARJETA DE CREDITO", "ZELLE TC", "TC - MO"}

GROUP_ALIASES = {
    "INBOUND AM": "INBOUND AM",
    "INBOUND AM(475)": "INBOUND AM",
    "INBOUND AM (475)": "INBOUND AM",
    "INBOUND PM": "INBOUND PM",
    "INBOUND PM(479)": "INBOUND PM",
    "INBOUND PM (479)": "INBOUND PM",
    "REDES SOCIALES AM": "REDES SOCIALES AM",
    "REDES SOCIALES AM(1485)": "REDES SOCIALES AM",
    "REDES SOCIALES AM (1485)": "REDES SOCIALES AM",
    "REDES SOCIALES PM": "REDES SOCIALES PM",
    "REDES SOCIALES PM(1486)": "REDES SOCIALES PM",
    "REDES SOCIALES PM (1486)": "REDES SOCIALES PM",
    "REDES INGLES": "REDES INGLES",
    "REDES INGLÉS": "REDES INGLES",
    "REDES INGLES(1503)": "REDES INGLES",
    "REDES INGLES (1503)": "REDES INGLES",
    "REDES INGLÉS (1503)": "REDES INGLES",
    "REDES INGLÉS(1503)": "REDES INGLES",
    "REORDEN AM": "BARRANQUILLA AM",
    "REORDEN AM(918)": "BARRANQUILLA AM",
    "REORDEN AM (918)": "BARRANQUILLA AM",
    "REORDEN PM": "BARRANQUILLA PM",
    "REORDEN PM(1944)": "BARRANQUILLA PM",
    "REORDEN PM (1944)": "BARRANQUILLA PM",
}

SHEET_NAME_BY_GROUP = {
    "INBOUND AM": "INBOUND AM",
    "INBOUND PM": "INBOUND PM",
    "REDES INGLES": "REDES INGLÉS",
    "REDES SOCIALES AM": "REDES SOCIALES AM",
    "REDES SOCIALES PM": "REDES SOCIALES PM",
}

OUTPUT_GROUP_ORDER = [
    "INBOUND AM",
    "INBOUND PM",
    "REDES INGLES",
    "REDES SOCIALES AM",
    "REDES SOCIALES PM",
]

FIXED_RULES = {
    "inbound": {
        "amount_band": {"min": 2450, "max": 2900, "rate": 0.01},
        "tier_rates": {"le_219": 0.015, "gt_219_le_269": 0.02, "gt_269": 0.025},
        "bonus_recaudo": [
            (3000, 13), (3250, 18), (3500, 29), (3750, 32), (4000, 37), (4250, 40), (4500, 44),
            (4750, 47), (5000, 50), (5250, 53), (5500, 56), (5550, 60), (5750, 65), (6000, 70),
            (6250, 75), (6500, 80), (6750, 85), (7000, 90), (7250, 95), (7500, 100), (7750, 105), (8000, 110),
        ],
        "bonus_rank_recaudo": [25, 20, 12],
        "bonus_rank_despacho": [25, 20, 12],
        "bonus_rank_ppv": [40, 25, 20, 10, 5],
        "bonus_rank_digital": [25, 20, 12],
    },
    "social": {
        "base_per_sale": 2.0,
        "tier_rates": {"le_185": 2.75, "gt_185_le_219": 5.0, "gt_219": 6.0},
        "bonus_pagados": [(11, 8), (12, 15), (14, 19), (15, 27), (16, 31), (17, 35), (18, 38), (19, 41), (20, 45), (22, 50), (24, 54), (26, 58), (28, 63), (30, 68), (32, 73), (34, 78), (36, 81)],
        "bonus_9": 4,
        "bonus_10": 5,
        "bonus_rank_entregas": [25, 12, 8],
        "bonus_rank_despacho": [25, 12, 8],
        "bonus_rank_ppv": [40, 20, 15, 10, 8],
        "bonus_rank_tc": [25, 20, 15, 10, 8],
        "bonus_tc_exact_6": 4,
    },
    "english": {
        "base_rate": 0.015,
        "tier_rates": {"le_229": 0.025, "gt_229_le_259": 0.03, "gt_259": 0.035},
        "bonus_pagados": [(10, 20), (11, 23), (12, 25), (13, 27), (14, 29), (15, 31), (16, 33), (17, 35), (18, 38), (19, 41), (20, 45), (22, 50), (24, 54), (26, 58), (28, 63), (30, 68), (32, 73), (34, 78), (36, 81)],
        "bonus_8": 4,
        "bonus_9": 5,
        "bonus_rank_entregas": [15, 10, 8],
        "bonus_rank_ppv": [40, 20, 10],
        "bonus_rank_tc": [15, 10, 8],
    },
}


def clean_text(value) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalize_group(value) -> str:
    text = clean_text(value).upper()
    text = re.sub(r"\s+", " ", text)
    return GROUP_ALIASES.get(text, text)


def normalize_user(value) -> str:
    return clean_text(value).upper().replace(" ", "")


def parse_vendor_name(value: str) -> Tuple[str, Optional[str]]:
    text = clean_text(value)
    text = re.sub(r"\s+", " ", text)
    m = re.match(r"^(.*?)(?:\s*\((\d+)\)|-(.*))$", text)
    if m:
        name = clean_text(m.group(1))
        ident = m.group(2) or m.group(3)
        return name, ident.strip() if ident else None
    return text, None


def floor_lookup(value: float, thresholds: List[Tuple[float, float]], strict: bool = False) -> float:
    result = 0.0
    for threshold, payout in sorted(thresholds, key=lambda x: x[0]):
        if (value > threshold) if strict else (value >= threshold):
            result = payout
    return result


def rank_bonus(df: pd.DataFrame, metric_col: str, eligible_mask: pd.Series, bonuses: List[float]) -> Dict[Tuple[str, str], float]:
    bonus_map: Dict[Tuple[str, str], float] = {}
    eligible = df[eligible_mask].copy()
    if eligible.empty:
        return bonus_map
    for group, grp in eligible.groupby("group"):
        grp = grp.sort_values([metric_col, "vendor"], ascending=[False, True]).reset_index(drop=True)
        for idx, (_, row) in enumerate(grp.iterrows()):
            if idx >= len(bonuses):
                break
            bonus_map[(group, row["user"])] = float(bonuses[idx])
    return bonus_map


def _read_html_tables(path: Path) -> List[pd.DataFrame]:
    return pd.read_html(str(path), header=None)


def _normalize_header_df(df: pd.DataFrame, header_row_idx: int = 1) -> pd.DataFrame:
    df = df.copy()
    header = [clean_text(x) for x in df.iloc[header_row_idx].tolist()]
    out = df.iloc[header_row_idx + 1:].copy().reset_index(drop=True)
    out.columns = header
    return out


def read_cobros_report(path: str | Path) -> Dict[str, pd.DataFrame]:
    path = Path(path)
    tables = _read_html_tables(path)
    meta = _normalize_header_df(pd.concat([tables[0].T], axis=0), 0)
    resumen = _normalize_header_df(tables[1], 1)
    detalle = _normalize_header_df(tables[4], 1)
    total_cobros = _normalize_header_df(tables[5], 1)
    ordenes = _normalize_header_df(tables[6], 1)
    return {
        "meta": meta,
        "resumen": resumen,
        "detalle": detalle,
        "total_cobros": total_cobros,
        "ordenes": ordenes,
    }


def read_analisis_report(path: str | Path) -> pd.DataFrame:
    tables = pd.read_html(str(path), header=[0, 1, 2])
    df = tables[0].copy()
    df.columns = ["|".join([clean_text(x) for x in col if clean_text(x)]) for col in df.columns.to_flat_index()]
    df = df.rename(columns={col: "Id" for col in df.columns if col.startswith("Id|Id")})
    df = df.rename(columns={col: "Nombre" for col in df.columns if col.startswith("Nombre|Nombre")})
    df = df.rename(columns={col: "Ordenes" for col in df.columns if col.startswith("Ordenes|Ordenes")})
    df = df[["Id", "Nombre", "Ordenes"]].copy()
    df["seller_id"] = df["Id"].astype(str).str.extract(r"(\d+)")
    df["Nombre"] = df["Nombre"].map(clean_text)
    df["Ordenes"] = pd.to_numeric(df["Ordenes"], errors="coerce").fillna(0).astype(int)
    return df[["seller_id", "Nombre", "Ordenes"]]


def read_bnt_report(path: str | Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [clean_text(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(3, ws.max_row + 1):
        vals = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        rows.append(vals)
    return pd.DataFrame(rows)


@dataclass
class ProcessResult:
    detail_tables: Dict[str, pd.DataFrame]
    validations: pd.DataFrame
    total_table: pd.DataFrame
    penalties_table: pd.DataFrame
    date_start: str
    date_end: str


class CommissionProcessor:
    def __init__(self, cobros_path, analisis_path=None, bnt_clpg_path=None, bnt_dgtl_path=None, penalties_df: Optional[pd.DataFrame] = None):
        self.cobros_path = Path(cobros_path)
        self.analisis_path = Path(analisis_path) if analisis_path else None
        self.bnt_clpg_path = Path(bnt_clpg_path) if bnt_clpg_path else None
        self.bnt_dgtl_path = Path(bnt_dgtl_path) if bnt_dgtl_path else None
        self.penalties_input = penalties_df if penalties_df is not None else pd.DataFrame(columns=["group", "user", "vendor", "penalty_count", "penalty_discount", "penalty_note"])

    def load(self):
        reports = read_cobros_report(self.cobros_path)
        self.meta = reports["meta"]
        self.resumen = reports["resumen"]
        self.detalle = reports["detalle"]
        self.total_cobros = reports["total_cobros"]
        self.ordenes = reports["ordenes"]
        self.analisis = read_analisis_report(self.analisis_path) if self.analisis_path else pd.DataFrame(columns=["seller_id", "Nombre", "Ordenes"])
        self.bnt_clpg = read_bnt_report(self.bnt_clpg_path) if self.bnt_clpg_path else pd.DataFrame()
        self.bnt_dgtl = read_bnt_report(self.bnt_dgtl_path) if self.bnt_dgtl_path else pd.DataFrame()

    def _prepare_summary(self) -> pd.DataFrame:
        df = self.resumen.copy()
        df["vendor"] = df["Vendedor"].map(lambda x: parse_vendor_name(x)[0])
        df["seller_id"] = df["Vendedor"].map(lambda x: parse_vendor_name(x)[1])
        df["user"] = df["User"].map(normalize_user)
        df["group"] = df["Grupo (id)"].map(normalize_group)
        df["raw_qty"] = pd.to_numeric(df["Cantidad de cobros"], errors="coerce").fillna(0).astype(int)
        df["raw_amount"] = pd.to_numeric(df["Cobrado"], errors="coerce").fillna(0.0)
        df["listado_qty"] = pd.to_numeric(df.get("Listado", 0), errors="coerce").fillna(0).astype(int)
        df = df[df["group"].isin(set(OUTPUT_GROUP_ORDER))].copy()
        df = df[df["vendor"].astype(str).str.strip() != ""].copy()
        return df[["vendor", "seller_id", "user", "group", "raw_qty", "raw_amount", "listado_qty"]]

    def _prepare_detail_orders(self) -> pd.DataFrame:
        df = self.detalle.copy()
        df["vendor"] = df["Vendedor"].map(lambda x: parse_vendor_name(x)[0])
        df["seller_id"] = df["Vendedor"].map(lambda x: parse_vendor_name(x)[1])
        df["user"] = df["User"].map(normalize_user)
        df["group"] = df["Grupo (id)"].map(normalize_group)
        df["order_id"] = df["Id. Orden"].astype(str).str.extract(r"(\d+)")
        df["payment_method"] = df["Forma de Pago"].map(lambda x: clean_text(x).upper())
        df["credit"] = pd.to_numeric(df["Credito"], errors="coerce").fillna(0.0)
        df["debit"] = pd.to_numeric(df["Debito"], errors="coerce").fillna(0.0)
        df["amount"] = df["credit"] - df["debit"]
        grouped = (
            df.groupby(["vendor", "seller_id", "user", "group", "order_id"], dropna=False)
            .agg(
                amount=("amount", "sum"),
                payment_method=("payment_method", lambda s: ";".join(sorted(set([x for x in s if x])))),
            )
            .reset_index()
        )
        grouped = grouped[grouped["amount"] > 0].copy()
        grouped = grouped[grouped["group"].isin(set(OUTPUT_GROUP_ORDER))].copy()
        grouped["tc_flag"] = grouped["payment_method"].map(lambda s: int(any(x in TC_PAYMENT_METHODS for x in s.split(";"))))
        grouped["source"] = "cobros"
        return grouped

    def _prepare_claim_orders(self) -> pd.DataFrame:
        empty = pd.DataFrame(columns=["vendor", "seller_id", "user", "group", "order_id", "amount", "payment_method", "tc_flag", "source"])
        if self.bnt_clpg.empty:
            return empty
        df = self.bnt_clpg.copy()
        df["group"] = df.get("Grupo", df.get("User Group", "")).map(normalize_group)
        df = df[df["group"].isin(["REDES SOCIALES AM", "REDES SOCIALES PM"])].copy()
        if df.empty:
            return empty
        df["vendor"] = df["Vendedor"].map(lambda x: parse_vendor_name(x)[0])
        df["seller_id"] = df["Vendedor"].map(lambda x: parse_vendor_name(x)[1])
        df["user"] = df.get("User", "").map(normalize_user)
        df["order_id"] = df["Id Orden"].astype(str).str.extract(r"(\d+)")
        df["amount"] = pd.to_numeric(df.get("Monto", 0), errors="coerce").fillna(0.0)
        payment_col = "Payment"
        if payment_col not in df.columns:
            df[payment_col] = ""
        df["payment_method"] = df[payment_col].map(lambda x: clean_text(x).upper())
        df["tc_flag"] = df["payment_method"].map(lambda s: int(s in TC_PAYMENT_METHODS))
        df["source"] = "claim"
        return df[["vendor", "seller_id", "user", "group", "order_id", "amount", "payment_method", "tc_flag", "source"]].copy()

    def _prepare_dispatches(self, seller_map: pd.DataFrame) -> pd.DataFrame:
        if self.analisis.empty:
            return pd.DataFrame(columns=["seller_id", "vendor", "user", "group", "dispatch_orders"])
        m = seller_map[["seller_id", "vendor", "user", "group"]].drop_duplicates()
        df = self.analisis.merge(m, on="seller_id", how="left")
        df["vendor"] = df["vendor"].fillna(df["Nombre"])
        df["dispatch_orders"] = pd.to_numeric(df["Ordenes"], errors="coerce").fillna(0).astype(int)
        return df[["seller_id", "vendor", "user", "group", "dispatch_orders"]]

    def _prepare_digitals(self) -> pd.DataFrame:
        if self.bnt_dgtl.empty:
            return pd.DataFrame(columns=["user", "group", "digital_orders"])
        df = self.bnt_dgtl.copy()
        df["group"] = df.get("Grupo", df.get("User Group", "")).map(normalize_group)
        df["user"] = df.get("User", "").map(normalize_user)
        df["status"] = df["Status"].map(lambda x: clean_text(x).upper())
        df = df[df["group"].isin(["INBOUND AM", "INBOUND PM"])].copy()
        df = df[df["status"] != "CANCELADA"].copy()
        return df.groupby(["user", "group"], as_index=False).size().rename(columns={"size": "digital_orders"})

    def _prepare_penalties(self) -> pd.DataFrame:
        df = self.penalties_input.copy()
        if df.empty:
            return pd.DataFrame(columns=["group", "user", "vendor", "penalty_count", "penalty_discount", "penalty_note"])
        expected = ["group", "user", "vendor", "penalty_count", "penalty_discount", "penalty_note"]
        for col in expected:
            if col not in df.columns:
                df[col] = "" if col in {"group", "user", "vendor", "penalty_note"} else 0
        df["group"] = df["group"].map(normalize_group)
        df["user"] = df["user"].map(normalize_user)
        df["vendor"] = df["vendor"].map(clean_text)
        df["penalty_count"] = pd.to_numeric(df["penalty_count"], errors="coerce").fillna(0)
        df["penalty_discount"] = pd.to_numeric(df["penalty_discount"], errors="coerce").fillna(0.0)
        df["penalty_note"] = df["penalty_note"].map(clean_text)
        df = df[df["group"].isin(set(OUTPUT_GROUP_ORDER))].copy()
        return df.groupby(["group", "user", "vendor"], as_index=False).agg(
            penalty_count=("penalty_count", "sum"),
            penalty_discount=("penalty_discount", "sum"),
            penalty_note=("penalty_note", lambda s: " | ".join([x for x in s if x])),
        )

    def _build_base(self) -> pd.DataFrame:
        summary = self._prepare_summary()
        detail_orders = self._prepare_detail_orders()
        claim_orders = self._prepare_claim_orders()
        orders_all = pd.concat([detail_orders, claim_orders], ignore_index=True, sort=False)

        order_metrics = (
            orders_all.groupby(["vendor", "seller_id", "user", "group"], dropna=False)
            .agg(
                validated_qty=("order_id", "nunique"),
                validated_amount=("amount", "sum"),
                tc_count=("tc_flag", "sum"),
                duplicate_source_rows=("source", lambda s: int((s == "claim").sum())),
            )
            .reset_index()
        )

        def bucket_metrics(group_df: pd.DataFrame) -> Dict[str, float]:
            g = clean_text(group_df["group"].iloc[0]).upper()
            amounts = group_df["amount"].tolist()
            if g.startswith("INBOUND"):
                return {
                    "amt_le_219": sum(v for v in amounts if v <= 219),
                    "amt_219_269": sum(v for v in amounts if 219 < v <= 269),
                    "amt_gt_269": sum(v for v in amounts if v > 269),
                    "cnt_le_185": 0,
                    "cnt_185_219": 0,
                    "cnt_gt_219": 0,
                    "cnt_le_229": 0,
                    "cnt_229_259": 0,
                    "cnt_gt_259": 0,
                }
            if g.startswith("REDES SOCIALES"):
                return {
                    "amt_le_219": 0,
                    "amt_219_269": 0,
                    "amt_gt_269": 0,
                    "cnt_le_185": sum(1 for v in amounts if v <= 185),
                    "cnt_185_219": sum(1 for v in amounts if 185 < v <= 219),
                    "cnt_gt_219": sum(1 for v in amounts if v > 219),
                    "cnt_le_229": 0,
                    "cnt_229_259": 0,
                    "cnt_gt_259": 0,
                }
            if g == "REDES INGLES":
                return {
                    "amt_le_219": 0,
                    "amt_219_269": 0,
                    "amt_gt_269": 0,
                    "cnt_le_185": 0,
                    "cnt_185_219": 0,
                    "cnt_gt_219": 0,
                    "cnt_le_229": sum(1 for v in amounts if v <= 229),
                    "cnt_229_259": sum(1 for v in amounts if 229 < v <= 259),
                    "cnt_gt_259": sum(1 for v in amounts if v > 259),
                }
            return {k: 0 for k in ["amt_le_219", "amt_219_269", "amt_gt_269", "cnt_le_185", "cnt_185_219", "cnt_gt_219", "cnt_le_229", "cnt_229_259", "cnt_gt_259"]}

        bucket_rows = []
        for _, grp in orders_all.groupby(["vendor", "seller_id", "user", "group"], dropna=False):
            data = bucket_metrics(grp)
            data.update({
                "vendor": grp["vendor"].iloc[0],
                "seller_id": grp["seller_id"].iloc[0],
                "user": grp["user"].iloc[0],
                "group": grp["group"].iloc[0],
            })
            bucket_rows.append(data)
        bucket_df = pd.DataFrame(bucket_rows)

        seller_map = pd.concat([
            summary[["seller_id", "vendor", "user", "group"]],
            order_metrics[["seller_id", "vendor", "user", "group"]],
        ], ignore_index=True).drop_duplicates()
        seller_map = seller_map[seller_map["group"].isin(set(OUTPUT_GROUP_ORDER))].copy()
        seller_map = seller_map[seller_map["vendor"].astype(str).str.strip() != ""].copy()

        dispatch_df = self._prepare_dispatches(seller_map)
        digital_df = self._prepare_digitals()
        penalties_df = self._prepare_penalties()

        base = seller_map.merge(summary, on=["seller_id", "vendor", "user", "group"], how="left")
        base = base.merge(order_metrics, on=["seller_id", "vendor", "user", "group"], how="left")
        if not bucket_df.empty:
            base = base.merge(bucket_df, on=["seller_id", "vendor", "user", "group"], how="left")
        else:
            for col in ["amt_le_219", "amt_219_269", "amt_gt_269", "cnt_le_185", "cnt_185_219", "cnt_gt_219", "cnt_le_229", "cnt_229_259", "cnt_gt_259"]:
                base[col] = 0
        base = base.merge(dispatch_df[["seller_id", "dispatch_orders"]], on="seller_id", how="left")
        base = base.merge(digital_df, on=["user", "group"], how="left")
        base = base.merge(penalties_df[["group", "user", "vendor", "penalty_count", "penalty_discount", "penalty_note"]], on=["group", "user", "vendor"], how="left")

        for col in [
            "raw_qty", "raw_amount", "listado_qty", "validated_qty", "validated_amount", "tc_count", "duplicate_source_rows",
            "dispatch_orders", "digital_orders", "amt_le_219", "amt_219_269", "amt_gt_269", "cnt_le_185", "cnt_185_219", "cnt_gt_219",
            "cnt_le_229", "cnt_229_259", "cnt_gt_259", "penalty_count", "penalty_discount",
        ]:
            if col not in base.columns:
                base[col] = 0
            base[col] = pd.to_numeric(base[col], errors="coerce").fillna(0)
        if "penalty_note" not in base.columns:
            base["penalty_note"] = ""

        base["quantity"] = base["validated_qty"].astype(int)
        base["amount"] = base.apply(
            lambda r: float(r["raw_amount"]) if r["group"] in ["INBOUND AM", "INBOUND PM"] and r["raw_amount"] > 0 else float(r["validated_amount"]),
            axis=1,
        )
        base["ppv"] = base.apply(lambda r: (r["amount"] / r["quantity"]) if r["quantity"] else 0, axis=1)
        base = base.sort_values(["group", "amount", "vendor"], ascending=[True, False, True]).reset_index(drop=True)
        self.orders_all = orders_all
        return base

    def _apply_rules(self, base: pd.DataFrame) -> pd.DataFrame:
        df = base.copy()
        for col in ["commission_usd", "bonus_payments", "bonus_rank_1", "bonus_rank_2", "bonus_ppv", "bonus_tc", "bonus_digital", "total_usd", "total_pesos"]:
            df[col] = 0.0

        inbound_rules = FIXED_RULES["inbound"]
        inbound_mask = df["group"].isin(["INBOUND AM", "INBOUND PM"])
        band_mask = inbound_mask & df["amount"].between(inbound_rules["amount_band"]["min"], inbound_rules["amount_band"]["max"], inclusive="both")
        df.loc[band_mask, "commission_usd"] = df.loc[band_mask, "amount"] * inbound_rules["amount_band"]["rate"]
        inbound_tier_mask = inbound_mask & (df["amount"] > inbound_rules["amount_band"]["max"])
        df.loc[inbound_tier_mask, "commission_usd"] = (
            df.loc[inbound_tier_mask, "amt_le_219"] * inbound_rules["tier_rates"]["le_219"]
            + df.loc[inbound_tier_mask, "amt_219_269"] * inbound_rules["tier_rates"]["gt_219_le_269"]
            + df.loc[inbound_tier_mask, "amt_gt_269"] * inbound_rules["tier_rates"]["gt_269"]
        )
        df.loc[inbound_mask, "bonus_payments"] = df.loc[inbound_mask, "amount"].map(lambda x: floor_lookup(x, inbound_rules["bonus_recaudo"], strict=True))
        bonus_recaudo = rank_bonus(df, "amount", inbound_mask & (df["amount"] > 3300), inbound_rules["bonus_rank_recaudo"])
        bonus_desp = rank_bonus(df, "dispatch_orders", inbound_mask & (df["amount"] > 3300) & (df["dispatch_orders"] > 0), inbound_rules["bonus_rank_despacho"])
        bonus_ppv = rank_bonus(df, "ppv", inbound_mask & (df["amount"] >= 3500) & (df["ppv"] > 209), inbound_rules["bonus_rank_ppv"])
        bonus_digital = rank_bonus(df, "digital_orders", inbound_mask & (df["amount"] > 3300) & (df["digital_orders"] >= 5), inbound_rules["bonus_rank_digital"])

        social_rules = FIXED_RULES["social"]
        social_mask = df["group"].isin(["REDES SOCIALES AM", "REDES SOCIALES PM"])
        small_social = social_mask & (df["quantity"] <= 10)
        large_social = social_mask & (df["quantity"] >= 11)
        df.loc[small_social, "commission_usd"] = df.loc[small_social, "quantity"] * social_rules["base_per_sale"]
        df.loc[large_social, "commission_usd"] = (
            df.loc[large_social, "cnt_le_185"] * social_rules["tier_rates"]["le_185"]
            + df.loc[large_social, "cnt_185_219"] * social_rules["tier_rates"]["gt_185_le_219"]
            + df.loc[large_social, "cnt_gt_219"] * social_rules["tier_rates"]["gt_219"]
        )
        df.loc[social_mask, "bonus_payments"] = df.loc[social_mask, "quantity"].map(lambda q: social_rules["bonus_9"] if q == 9 else (social_rules["bonus_10"] if q == 10 else floor_lookup(q, social_rules["bonus_pagados"])))
        social_ent = rank_bonus(df, "quantity", social_mask & (df["quantity"] > 11), social_rules["bonus_rank_entregas"])
        social_desp = rank_bonus(df, "dispatch_orders", social_mask & (df["dispatch_orders"] > 11), social_rules["bonus_rank_despacho"])
        social_ppv = rank_bonus(df, "ppv", social_mask & (df["dispatch_orders"] >= 12) & (df["ppv"] >= 199), social_rules["bonus_rank_ppv"])
        social_tc = rank_bonus(df, "tc_count", social_mask & (df["quantity"] >= 12) & (df["tc_count"] > 6), social_rules["bonus_rank_tc"])
        exact6_mask = social_mask & (df["quantity"] >= 12) & (df["tc_count"] == 6)
        df.loc[exact6_mask, "bonus_tc"] = social_rules["bonus_tc_exact_6"]

        eng_rules = FIXED_RULES["english"]
        english_mask = df["group"] == "REDES INGLES"
        small_eng = english_mask & (df["quantity"] <= 9)
        large_eng = english_mask & (df["quantity"] >= 10)
        df.loc[small_eng, "commission_usd"] = df.loc[small_eng, "quantity"] * eng_rules["base_rate"]
        df.loc[large_eng, "commission_usd"] = (
            df.loc[large_eng, "cnt_le_229"] * eng_rules["tier_rates"]["le_229"]
            + df.loc[large_eng, "cnt_229_259"] * eng_rules["tier_rates"]["gt_229_le_259"]
            + df.loc[large_eng, "cnt_gt_259"] * eng_rules["tier_rates"]["gt_259"]
        )
        df.loc[english_mask, "bonus_payments"] = df.loc[english_mask, "quantity"].map(lambda q: eng_rules["bonus_8"] if q == 8 else (eng_rules["bonus_9"] if q == 9 else floor_lookup(q, eng_rules["bonus_pagados"])))
        eng_ent = rank_bonus(df, "quantity", english_mask & (df["quantity"] > 10), eng_rules["bonus_rank_entregas"])
        eng_ppv = rank_bonus(df, "ppv", english_mask & (df["quantity"] >= 10) & (df["ppv"] >= 259), eng_rules["bonus_rank_ppv"])
        eng_tc = rank_bonus(df, "tc_count", english_mask & (df["quantity"] >= 10) & (df["tc_count"] >= 5), eng_rules["bonus_rank_tc"])

        for mp, col in [
            (bonus_recaudo, "bonus_rank_1"),
            (bonus_desp, "bonus_rank_2"),
            (social_ent, "bonus_rank_1"),
            (social_desp, "bonus_rank_2"),
            (eng_ent, "bonus_rank_1"),
            (bonus_ppv, "bonus_ppv"),
            (bonus_digital, "bonus_digital"),
            (social_ppv, "bonus_ppv"),
            (social_tc, "bonus_tc"),
            (eng_ppv, "bonus_ppv"),
            (eng_tc, "bonus_tc"),
        ]:
            for (group, user), bonus in mp.items():
                idx = (df["group"] == group) & (df["user"] == user)
                df.loc[idx, col] = df.loc[idx, col] + bonus

        df["total_usd"] = (
            df["commission_usd"] + df["bonus_payments"] + df["bonus_rank_1"] + df["bonus_rank_2"] + df["bonus_ppv"] + df["bonus_tc"] + df["bonus_digital"] - df["penalty_discount"]
        )
        df["total_pesos"] = df["total_usd"] * TRM_PESOS
        return df

    def process(self) -> ProcessResult:
        self.load()
        base = self._build_base()
        calc = self._apply_rules(base)
        validations = self._build_validations(calc)
        detail_tables = self._build_detail_tables(calc)
        total_table = self._build_total_table(detail_tables)
        penalties_table = calc[["group", "vendor", "user", "penalty_count", "penalty_discount", "penalty_note"]].copy()
        penalties_table = penalties_table[(penalties_table["penalty_count"] > 0) | (penalties_table["penalty_discount"] > 0) | (penalties_table["penalty_note"].astype(str) != "")].reset_index(drop=True)
        self.calc = calc
        return ProcessResult(
            detail_tables=detail_tables,
            validations=validations,
            total_table=total_table,
            penalties_table=penalties_table,
            date_start=self._extract_meta_date("FECHA INICIO"),
            date_end=self._extract_meta_date("FECHA FIN"),
        )

    def _extract_meta_date(self, key: str) -> str:
        try:
            tables = _read_html_tables(self.cobros_path)
            meta = tables[0]
            for i in range(len(meta)):
                if clean_text(meta.iloc[i, 0]).upper() == key.upper():
                    return clean_text(meta.iloc[i, 1])
        except Exception:
            pass
        return ""

    def _build_validations(self, calc: pd.DataFrame) -> pd.DataFrame:
        df = calc.copy()
        df["qty_diff"] = df["raw_qty"] - df["quantity"]
        df["missing_analysis"] = df["dispatch_orders"].eq(0) & df["group"].isin(["INBOUND AM", "INBOUND PM", "REDES SOCIALES AM", "REDES SOCIALES PM"])
        out = df[["group", "vendor", "user", "raw_qty", "quantity", "qty_diff", "raw_amount", "amount", "dispatch_orders", "digital_orders", "duplicate_source_rows", "penalty_count", "penalty_discount", "penalty_note", "missing_analysis"]].copy()
        return out.sort_values(["group", "qty_diff", "vendor"], ascending=[True, False, True])

    def _build_detail_tables(self, calc: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        tables: Dict[str, pd.DataFrame] = {}
        for group in OUTPUT_GROUP_ORDER:
            grp = calc[calc["group"] == group].copy()
            if group in ["INBOUND AM", "INBOUND PM"]:
                cols = [
                    ("vendor", "AGENTE COMERCIAL"), ("user", "USER"), ("group", "GRUPO"), ("quantity", "CANTIDAD"), ("amount", "RECAUDO USD"),
                    ("amt_le_219", "VENTAS < $219"), ("amt_219_269", "VENTAS > $219"), ("amt_gt_269", "VENTAS > $269"), ("commission_usd", "COMISIÓN USD"),
                    ("bonus_payments", "BONO RECAUDO"), ("bonus_rank_1", "BONO > REC"), ("dispatch_orders", "VENTAS DESP"), ("bonus_rank_2", "BONO > DESP"),
                    ("ppv", "PPV"), ("bonus_ppv", "BONO PPV"), ("digital_orders", "VENTAS DIGITALES"), ("bonus_digital", "BONO VNT. DIG."),
                    ("penalty_count", "PENALD."), ("penalty_discount", "DESC PENALD."), ("total_usd", "TOTAL USD"), ("total_pesos", "TOTAL PESOS"),
                ]
            elif group in ["REDES SOCIALES AM", "REDES SOCIALES PM"]:
                cols = [
                    ("vendor", "AGENTE COMERCIAL"), ("user", "USER"), ("group", "GRUPO"), ("quantity", "CANTIDAD"), ("cnt_le_185", "VENTAS < $185"),
                    ("cnt_185_219", "VENTAS < $219"), ("cnt_gt_219", "VENTAS > $219"), ("commission_usd", "COMISIÓN USD"), ("bonus_payments", "BONO PAGOS"),
                    ("bonus_rank_1", "BONO > ENT"), ("dispatch_orders", "VENTAS DESP"), ("bonus_rank_2", "BONO > DESP"), ("ppv", "PPV"),
                    ("bonus_ppv", "BONO PPV"), ("tc_count", "CANT T.C"), ("bonus_tc", "BONO T.C"), ("penalty_count", "PENALD."),
                    ("penalty_discount", "DESC PENALD."), ("total_usd", "TOTAL USD"), ("total_pesos", "TOTAL PESOS"),
                ]
            else:
                cols = [
                    ("vendor", "AGENTE COMERCIAL"), ("user", "USER"), ("group", "GRUPO"), ("quantity", "CANTIDAD"), ("amount", "RECAUDO USD"),
                    ("cnt_le_229", "VENTAS < $229"), ("cnt_229_259", "VENTAS < $259"), ("cnt_gt_259", "VENTAS > $259"), ("commission_usd", "COMISIÓN USD"),
                    ("bonus_payments", "BONO PAGOS"), ("bonus_rank_1", "BONO > ENT"), ("ppv", "PPV"), ("bonus_ppv", "BONO PPV"),
                    ("tc_count", "CANT T.C"), ("bonus_tc", "BONO T.C"), ("penalty_count", "PENALD."), ("penalty_discount", "DESC PENALD."),
                    ("total_usd", "TOTAL USD"), ("total_pesos", "TOTAL PESOS"),
                ]
            tbl = grp[[c[0] for c in cols]].copy() if not grp.empty else pd.DataFrame(columns=[c[1] for c in cols])
            tbl.columns = [c[1] for c in cols]
            if not tbl.empty:
                tbl = tbl.sort_values("TOTAL PESOS", ascending=False).reset_index(drop=True)
            tables[group] = tbl
        return tables

    def _build_total_table(self, tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        rows = []
        for group in OUTPUT_GROUP_ORDER:
            df = tables.get(group, pd.DataFrame())
            total = pd.to_numeric(df.get("TOTAL PESOS", pd.Series(dtype=float)), errors="coerce").fillna(0).sum() if not df.empty else 0.0
            rows.append({"GRUPO": group, "TOTAL": total, "CAMBIO PORCENTUAL SEMANA ANTERIOR": None})
        total_df = pd.DataFrame(rows)
        total_grupos = total_df["TOTAL"].sum()
        return pd.concat([total_df, pd.DataFrame([
            {"GRUPO": "TOTAL GRUPOS", "TOTAL": total_grupos, "CAMBIO PORCENTUAL SEMANA ANTERIOR": None},
            {"GRUPO": "TOTAL CAMPAÑAS VERDADERO", "TOTAL": total_grupos, "CAMBIO PORCENTUAL SEMANA ANTERIOR": None},
        ])], ignore_index=True)

    def export_to_template(self, template_path: str | Path, output_path: str | Path, result: Optional[ProcessResult] = None):
        if result is None:
            result = self.process()
        wb = load_workbook(template_path)
        total_ws = wb["TOTAL "] if "TOTAL " in wb.sheetnames else wb["TOTAL"]
        total_ws["B2"] = f"LIQUIDACIÓN COMISIONES SEMANALES\n{result.date_start} AL {result.date_end}".strip()
        total_map = {row["GRUPO"]: row["TOTAL"] for _, row in result.total_table.iterrows()}
        target_rows = {"INBOUND AM": 5, "INBOUND PM": 6, "REDES INGLES": 7, "REDES SOCIALES AM": 8, "REDES SOCIALES PM": 9, "BARRANQUILLA AM": 10, "BARRANQUILLA PM": 11, "TOTAL GRUPOS": 12, "TOTAL CAMPAÑAS VERDADERO": 13}
        labels = {"INBOUND AM": "IN BOUND AM", "INBOUND PM": "IN BOUND PM", "REDES INGLES": "REDES INGLÉS", "REDES SOCIALES AM": "REDES SOCIALES AM", "REDES SOCIALES PM": "REDES SOCIALES PM", "BARRANQUILLA AM": "BARRANQUILLA AM", "BARRANQUILLA PM": "BARRANQUILLA PM", "TOTAL GRUPOS": "TOTAL GRUPOS", "TOTAL CAMPAÑAS VERDADERO": "TOTAL CAMPAÑAS VERDADERO"}
        for key, row_num in target_rows.items():
            total_ws[f"B{row_num}"] = labels[key]
            total_ws[f"C{row_num}"] = float(total_map.get(key, 0))

        for group, table in result.detail_tables.items():
            sheet_name = SHEET_NAME_BY_GROUP[group]
            self._write_group_sheet_exact(wb[sheet_name], sheet_name, table)

        wb.save(output_path)

    def _write_group_sheet_exact(self, ws, sheet_name: str, table: pd.DataFrame):
        header_row = 8
        data_start = 9
        cfg = self._sheet_config(sheet_name)
        original_total_row = self._find_total_row(ws, cfg["last_col"])
        if original_total_row is None:
            original_total_row = ws.max_row
        existing_capacity = max(0, original_total_row - data_start)
        needed_rows = len(table)
        if needed_rows > existing_capacity:
            ws.insert_rows(original_total_row, amount=needed_rows - existing_capacity)
            for i in range(existing_capacity, needed_rows):
                self._copy_row_format(ws, original_total_row - 1, data_start + i)
        elif needed_rows < existing_capacity:
            blank_start = data_start + needed_rows
            blank_end = data_start + existing_capacity - 1
            self._clear_rows(ws, blank_start, blank_end)

        total_row = data_start + needed_rows if needed_rows > 0 else data_start
        if total_row != original_total_row:
            self._copy_row_format(ws, original_total_row, total_row)
            self._copy_row_format(ws, min(original_total_row + 1, ws.max_row), total_row + 1)
            self._clear_row_values(ws, total_row)
            self._clear_row_values(ws, total_row + 1)
            if original_total_row > total_row:
                self._clear_rows(ws, total_row + 1, original_total_row)

        for idx, (_, row) in enumerate(table.iterrows(), start=data_start):
            for col_idx, col_name in enumerate(cfg["columns"], start=2):
                value = self._display_group(row.get("GRUPO")) if col_name == "GRUPO" else row.get(col_name)
                ws.cell(idx, col_idx, self._excel_blank(value))

        if needed_rows == 0:
            for col_idx in range(2, cfg["last_col"] + 1):
                ws.cell(data_start, col_idx).value = None
        penalties_total = float(pd.to_numeric(table.get("DESC PENALD.", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not table.empty else 0.0
        grand_total = float(pd.to_numeric(table.get("TOTAL PESOS", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not table.empty else 0.0
        ws.cell(3, cfg["last_col"], penalties_total)
        self._reset_total_merge(ws, cfg["last_col"], total_row)
        ws.cell(total_row, cfg["last_col"], grand_total)

        # Clear any stale values below total row within used range
        if total_row + 1 <= ws.max_row:
            self._clear_rows(ws, total_row + 1, ws.max_row, start_col=2, end_col=cfg["last_col"])

    def _sheet_config(self, sheet_name: str) -> Dict[str, object]:
        if sheet_name in ["INBOUND AM", "INBOUND PM"]:
            return {
                "columns": ["AGENTE COMERCIAL", "USER", "GRUPO", "CANTIDAD", "RECAUDO USD", "VENTAS < $219", "VENTAS > $219", "VENTAS > $269", "COMISIÓN USD", "BONO RECAUDO", "BONO > REC", "VENTAS DESP", "BONO > DESP", "PPV", "BONO PPV", "VENTAS DIGITALES", "BONO VNT. DIG.", "PENALD.", "DESC PENALD.", "TOTAL USD", "TOTAL PESOS"],
                "last_col": 22,
            }
        if sheet_name in ["REDES SOCIALES AM", "REDES SOCIALES PM"]:
            return {
                "columns": ["AGENTE COMERCIAL", "USER", "GRUPO", "CANTIDAD", "VENTAS < $185", "VENTAS < $219", "VENTAS > $219", "COMISIÓN USD", "BONO PAGOS", "BONO > ENT", "VENTAS DESP", "BONO > DESP", "PPV", "BONO PPV", "CANT T.C", "BONO T.C", "PENALD.", "DESC PENALD.", "TOTAL USD", "TOTAL PESOS"],
                "last_col": 21,
            }
        return {
            "columns": ["AGENTE COMERCIAL", "USER", "GRUPO", "CANTIDAD", "RECAUDO USD", "VENTAS < $229", "VENTAS < $259", "VENTAS > $259", "COMISIÓN USD", "BONO PAGOS", "BONO > ENT", "PPV", "BONO PPV", "CANT T.C", "BONO T.C", "PENALD.", "DESC PENALD.", "TOTAL USD", "TOTAL PESOS"],
            "last_col": 20,
        }

    def _reset_total_merge(self, ws, last_col: int, total_row: int):
        col_letter = get_column_letter(last_col)
        for rng in list(ws.merged_cells.ranges):
            if rng.min_col == last_col and rng.max_col == last_col and rng.min_row >= 8:
                ws.unmerge_cells(str(rng))
        ws.merge_cells(f"{col_letter}{total_row}:{col_letter}{total_row + 1}")

    def _find_total_row(self, ws, last_col: int) -> Optional[int]:
        for r in range(ws.max_row, 0, -1):
            val = ws.cell(r, last_col).value
            if val not in (None, ""):
                return r
        return None

    def _copy_row_format(self, ws, src_row: int, dst_row: int):
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        for col in range(1, ws.max_column + 1):
            src = ws.cell(src_row, col)
            dst = ws.cell(dst_row, col)
            if src.has_style:
                dst._style = copy.copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.font:
                dst.font = copy.copy(src.font)
            if src.fill:
                dst.fill = copy.copy(src.fill)
            if src.border:
                dst.border = copy.copy(src.border)
            if src.alignment:
                dst.alignment = copy.copy(src.alignment)
            if src.protection:
                dst.protection = copy.copy(src.protection)

    def _clear_row_values(self, ws, row_idx: int, start_col: int = 2, end_col: Optional[int] = None):
        end_col = end_col or ws.max_column
        for col in range(start_col, end_col + 1):
            if not isinstance(ws.cell(row_idx, col), MergedCell):
                ws.cell(row_idx, col).value = None

    def _clear_rows(self, ws, start_row: int, end_row: int, start_col: int = 2, end_col: Optional[int] = None):
        if start_row > end_row:
            return
        end_col = end_col or ws.max_column
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    def _display_group(self, group_value: str) -> str:
        g = normalize_group(group_value)
        return {"INBOUND AM": "INBOUND AM (475)", "INBOUND PM": "INBOUND PM (479)", "REDES SOCIALES AM": "REDES SOCIALES AM (1485)", "REDES SOCIALES PM": "REDES SOCIALES PM (1486)", "REDES INGLES": "REDES INGLES (1503)"}.get(g, g)

    def _excel_blank(self, value):
        if value is None:
            return None
        try:
            if isinstance(value, float) and math.isnan(value):
                return None
        except Exception:
            pass
        return value
